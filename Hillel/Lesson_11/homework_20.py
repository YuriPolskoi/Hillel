# Homework 20 by Polskoi Yuri

import csv
import openpyxl

with open('hw_19.csv') as f:
    reader = csv.reader(f, delimiter=';')
    new_data = []
    for index, item in enumerate(reader):
        new_data.append(item[0])

data_list = [item.split(',') for item in new_data]

for lst in data_list:
    if len(lst) > 2:
        lst.pop(2)

work_book = openpyxl.Workbook()
work_book.create_sheet(title='Перший лист', index=0)
sheet = work_book['Перший лист']

for row_index, row in enumerate(data_list):
    for col_index, value in enumerate(row):
        cell = sheet.cell(row=row_index + 1, column=col_index + 1)
        cell.value = value

work_book.create_sheet(title='Другий лист', index=1)
sheet = work_book['Другий лист']

for row_index, row in enumerate(data_list):
    cell = sheet.cell(column=row_index + 1, row=1)
    cell.value = f"Person {row_index}" if row_index > 0 else ""
    for col_index, value in enumerate(row):
        cell = sheet.cell(column=row_index + 1, row=col_index + 2)
        cell.value = value

work_book.save('hw_20.xlsx')
